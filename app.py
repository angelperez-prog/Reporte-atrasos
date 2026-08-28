import imaplib
import email
from email.header import decode_header
import re
import unicodedata
from datetime import datetime, timedelta
import pandas as pd
import io
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill, Font
import streamlit as st
import plotly.express as px

st.set_page_config(page_title="Procesador de Atrasos", layout="wide")

st.title("Procesador de Reporte de Atrasos")
st.write("Ingresa tus credenciales para procesar el reporte y visualizar el análisis:")

IMAP_SERVER = "imap.gmail.com"

DIAS_SEMANA = {
    "Monday": "Lunes",
    "Tuesday": "Martes",
    "Wednesday": "Miércoles",
    "Thursday": "Jueves",
    "Friday": "Viernes",
    "Saturday": "Sábado",
    "Sunday": "Domingo"
}

def obtener_rango_15m(dt):
    minuto_inicio = (dt.minute // 15) * 15
    minuto_fin = minuto_inicio + 14
    return f"{dt.hour:02d}:{minuto_inicio:02d} - {dt.hour:02d}:{minuto_fin:02d}"

def limpiar_texto(texto):
    if not texto:
        return ""
    texto_limpio = unicodedata.normalize("NFKD", texto)
    return texto_limpio.replace('\xa0', ' ').replace('\u200b', '').strip()

def parse_email_datetime(dt_str):
    try:
        return datetime.strptime(dt_str.strip(), "%d/%m/%Y %H:%M:%S")
    except ValueError:
        return None

def decodificar_asunto(header_subject):
    if not header_subject:
        return ""
    decoded_list = decode_header(header_subject)
    subject_str = ""
    for bytes_or_str, encoding in decoded_list:
        if isinstance(bytes_or_str, bytes):
            subject_str += bytes_or_str.decode(encoding or 'utf-8', errors='ignore')
        else:
            subject_str += bytes_or_str
    return limpiar_texto(subject_str)

with st.form("login_form"):
    raw_email = st.text_input("Correo Institucional Completo", placeholder="usuario@estudiantes.soldelillimani.cl")
    raw_pass = st.text_input("Contraseña de Aplicación (16 letras)", type="password")
    submitted = st.form_submit_button("Generar y Procesar Excel")

if submitted:
    if not raw_email or not raw_pass:
        st.error("Por favor completa ambos campos.")
    else:
        EMAIL_ACCOUNT = unicodedata.normalize("NFKD", raw_email).encode('ascii', 'ignore').decode('utf-8').strip()
        PASSWORD = unicodedata.normalize("NFKD", raw_pass).encode('ascii', 'ignore').decode('utf-8').replace(' ', '').strip()

        try:
            status_text = st.empty()
            progress_bar = st.progress(0)

            status_text.text("Conectando al servidor IMAP...")
            mail = imaplib.IMAP4_SSL(IMAP_SERVER)
            mail.login(EMAIL_ACCOUNT, PASSWORD)
            mail.select("inbox")

            status_text.text("Buscando correos recibidos...")
            status, messages = mail.search(None, '(FROM "miguel.plaza@soldelillimani.cl")')
            
            if status != 'OK' or not messages[0]:
                status, messages = mail.search(None, 'SUBJECT "Atraso"')

            email_ids = messages[0].split()
            total_encontrados = len(email_ids)

            if total_encontrados == 0:
                st.warning("No se encontraron correos de atraso.")
            else:
                # Se procesan TODOS los correos encontrados sin límite
                email_ids_procesar = list(reversed(email_ids))
                total_procesar = len(email_ids_procesar)

                records = []

                regex_fechahora = re.compile(r"Fecha/Hora:\s*(\d{1,2}/\d{1,2}/\d{4}\s+\d{1,2}:\d{2}:\d{2})", re.IGNORECASE)
                regex_curso = re.compile(r"Curso:\s*([^\r\n]+)", re.IGNORECASE)
                regex_estimado = re.compile(r"Estimado/a\s+([\w\s,ÁÉÍÓÚÑáéíóúñ]+?),\s*(?:\r?\n|Informamos)", re.IGNORECASE)
                regex_asunto_nombre = re.compile(r"(?:Notificación de Atraso|Atraso)\s*:\s*([^\r\n]+)", re.IGNORECASE)
                regex_campo_alumno = re.compile(r"(?:Alumno|Estudiante|Nombre)\s*:\s*([^\r\n]+)", re.IGNORECASE)

                for idx, e_id in enumerate(email_ids_procesar):
                    status_text.text(f"Procesando correo {idx + 1} de {total_procesar}...")
                    progress_bar.progress((idx + 1) / total_procesar)

                    status, msg_data = mail.fetch(e_id, "(RFC822)")
                    if status != 'OK':
                        continue

                    for response_part in msg_data:
                        if isinstance(response_part, tuple):
                            msg = email.message_from_bytes(response_part[1])
                            asunto = decodificar_asunto(msg.get("Subject"))

                            body = ""
                            if msg.is_multipart():
                                for part in msg.walk():
                                    content_type = part.get_content_type()
                                    content_disposition = str(part.get("Content-Disposition"))
                                    if content_type == "text/plain" and "attachment" not in content_disposition:
                                        payload = part.get_payload(decode=True)
                                        charset = part.get_content_charset() or "utf-8"
                                        body = payload.decode(charset, errors="ignore")
                                        break
                            else:
                                payload = msg.get_payload(decode=True)
                                charset = msg.get_content_charset() or "utf-8"
                                body = payload.decode(charset, errors="ignore")

                            body = limpiar_texto(body)

                            match_asunto = regex_asunto_nombre.search(asunto)
                            match_estimado = regex_estimado.search(body)
                            match_campo = regex_campo_alumno.search(body)

                            if match_asunto:
                                nombre_alumno = match_asunto.group(1).strip()
                            elif match_campo:
                                nombre_alumno = match_campo.group(1).strip()
                            elif match_estimado:
                                nombre_alumno = match_estimado.group(1).strip()
                            else:
                                nombre_alumno = "No especificado"

                            match_c = regex_curso.search(body)
                            curso = match_c.group(1).strip() if match_c else "N/A"

                            match_fh = regex_fechahora.search(body)
                            if match_fh:
                                raw_dt = match_fh.group(1)
                                parsed_dt = parse_email_datetime(raw_dt)

                                if parsed_dt:
                                    fecha_str = parsed_dt.strftime("%Y-%m-%d")
                                    hora_str = parsed_dt.strftime("%H:%M:%S")

                                    dia_nombre = DIAS_SEMANA.get(parsed_dt.strftime("%A"), parsed_dt.strftime("%A"))
                                    bloque_15m = obtener_rango_15m(parsed_dt)

                                    lunes_semana = parsed_dt.date() - timedelta(days=parsed_dt.weekday())
                                    semana_str = f"Semana del {lunes_semana.strftime('%Y-%m-%d')}"

                                    limite = parsed_dt.replace(hour=9, minute=35, second=0, microsecond=0)
                                    estado = "Antes de las 9:35" if parsed_dt <= limite else "Después de las 9:35"

                                    records.append({
                                        "Alumno": nombre_alumno,
                                        "Curso": curso,
                                        "Fecha": fecha_str,
                                        "Día de la semana": dia_nombre,
                                        "Hora de registro": hora_str,
                                        "Estado": estado,
                                        "Rango horario (15m)": bloque_15m,
                                        "Semana": semana_str,
                                        "Mes": parsed_dt.strftime("%B")
                                    })

                mail.close()
                mail.logout()

                status_text.empty()
                progress_bar.empty()

                if records:
                    df = pd.DataFrame(records)
                    df.sort_values(by=["Fecha", "Hora de registro"], ascending=False, inplace=True)

                    st.success(f"¡Proceso finalizado! Registros procesados: {len(df)}")

                    # --- GENERACIÓN Y DESCARGA DE EXCEL ---
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name='Analisis_Atrasos')

                    output.seek(0)
                    wb = openpyxl.load_workbook(output)
                    ws = wb.active

                    for col in ws.columns:
                        max_len = max(len(str(cell.value or '')) for cell in col)
                        col_letter = get_column_letter(col[0].column)
                        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

                    fill_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    font_verde = Font(color="006100")
                    fill_rojo = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    font_rojo = Font(color="9C0006")

                    rango_estado = f"F2:F{len(df) + 1}"
                    ws.conditional_formatting.add(
                        rango_estado,
                        CellIsRule(operator='equal', formula=['"Antes de las 9:35"'], fill=fill_verde, font=font_verde)
                    )
                    ws.conditional_formatting.add(
                        rango_estado,
                        CellIsRule(operator='equal', formula=['"Después de las 9:35"'], fill=fill_rojo, font=font_rojo)
                    )

                    final_output = io.BytesIO()
                    wb.save(final_output)
                    excel_data = final_output.getvalue()

                    st.download_button(
                        label="📥 Descargar Excel de Atrasos Formateado",
                        data=excel_data,
                        file_name="analisis_atrasos.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                    # --- TABLA REGISTRO DETALLADO ---
                    st.subheader("📋 Registro Detallado")

                    def resaltar_fila(row):
                        if row["Estado"] == "Después de las 9:35":
                            return ['background-color: #ffc7ce; color: #9c0006; font-weight: bold'] * len(row)
                        elif row["Estado"] == "Antes de las 9:35":
                            return ['background-color: #c6efce; color: #006100'] * len(row)
                        return [''] * len(row)

                    st.dataframe(df.style.apply(resaltar_fila, axis=1), use_container_width=True)

                    st.markdown("---")

                    # --- GRÁFICOS Y ANÁLISIS ---
                    st.subheader("📊 Análisis Gráfico")

                    col_graf1, col_graf2 = st.columns(2)

                    with col_graf1:
                        # 1. Gráfico: Registros por semana (Barras)
                        st.markdown("### 1. Registros por semana")
                        dias_por_semana = df.groupby("Semana")["Fecha"].nunique()
                        promedio_dias_semana = dias_por_semana.mean()

                        st.metric(
                            label="Promedio de registros por semana",
                            value=f"{promedio_dias_semana:.1f} días / semana"
                        )

                        df_semana_dias = dias_por_semana.reset_index()
                        df_semana_dias.columns = ["Semana", "Días con Atraso"]

                        fig_semana = px.bar(
                            df_semana_dias,
                            x="Semana",
                            y="Días con Atraso",
                            text="Días con Atraso",
                            title="Registros por semana",
                            color="Días con Atraso",
                            color_continuous_scale="Blues"
                        )
                        fig_semana.update_traces(textposition="outside")
                        fig_semana.update_yaxes(dtick=1)
                        fig_semana.update_xaxes(type='category', tickmode='linear')
                        st.plotly_chart(fig_semana, use_container_width=True)

                    with col_graf2:
                        # 2. Gráfico: Distribución de hora de llegada
                        st.markdown("### 2. Distribución de hora de llegada")
                        
                        horas_dt = pd.to_datetime(df["Hora de registro"], format="%H:%M:%S")
                        promedio_segundos = (horas_dt.dt.hour * 3600 + horas_dt.dt.minute * 60 + horas_dt.dt.second).mean()
                        hora_promedio_str = str(timedelta(seconds=int(promedio_segundos)))

                        st.metric(
                            label="Hora Promedio de Llegada",
                            value=hora_promedio_str
                        )

                        df_rangos = df.groupby(["Rango horario (15m)", "Estado"]).size().reset_index(name="Cantidad")
                        df_rangos = df_rangos.sort_values(by="Rango horario (15m)")

                        fig_horas = px.bar(
                            df_rangos,
                            x="Rango horario (15m)",
                            y="Cantidad",
                            color="Estado",
                            title="Llegadas por bloques de 15 minutos",
                            color_discrete_map={
                                "Antes de las 9:35": "#2ca02c",
                                "Después de las 9:35": "#d62728"
                            }
                        )
                        st.plotly_chart(fig_horas, use_container_width=True)

                else:
                    st.warning("No se encontraron registros de correos procesables.")

        except Exception as e:
            st.error(f"Error al conectar o procesar: {e}")
